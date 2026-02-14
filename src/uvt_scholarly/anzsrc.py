# SPDX-FileCopyrightText: 2026 Alexandru Fikl <alexfikl@gmail.com>
# SPDX-License-Identifier: MIT

from __future__ import annotations

from typing import TYPE_CHECKING

from uvt_scholarly.logging import make_logger

if TYPE_CHECKING:
    import pathlib

log = make_logger(__name__)

# See:
ANZSRC_FOR_URL: str = "https://www.abs.gov.au/statistics/classifications/australian-and-new-zealand-standard-research-classification-anzsrc/2020/anzsrc2020_for.xlsx"
"""The full URL for the ANZSRC classification. The latest version can be found at
the [official website](https://www.abs.gov.au/statistics/classifications/australian-and-new-zealand-standard-research-classification-anzsrc/latest-release).
"""

ANZSRC_CLASSIFICATIONS: dict[int, str] = {
    30: "Agricultural, Veterinary and Food Sciences",
    3001: "Agricultural biotechnology",
    3002: "Agriculture, land and farm management",
    3003: "Animal production",
    3004: "Crop and pasture production",
    3005: "Fisheries sciences",
    3006: "Food sciences",
    3007: "Forestry sciences",
    3008: "Horticultural production",
    3009: "Veterinary sciences",
    3099: "Other agricultural, veterinary and food sciences",
    31: "Biological Sciences",
    3101: "Biochemistry and cell biology",
    3102: "Bioinformatics and computational biology",
    3103: "Ecology",
    3104: "Evolutionary biology",
    3105: "Genetics",
    3106: "Industrial biotechnology",
    3107: "Microbiology",
    3108: "Plant biology",
    3109: "Zoology",
    3199: "Other biological sciences",
    32: "Biomedical and Clinical Sciences",
    3201: "Cardiovascular medicine and haematology",
    3202: "Clinical sciences",
    3203: "Dentistry",
    3204: "Immunology",
    3205: "Medical biochemistry and metabolomics",
    3206: "Medical biotechnology",
    3207: "Medical microbiology",
    3208: "Medical physiology",
    3209: "Neurosciences",
    3210: "Nutrition and dietetics",
    3211: "Oncology and carcinogenesis",
    3212: "Ophthalmology and optometry",
    3213: "Paediatrics",
    3214: "Pharmacology and pharmaceutical sciences",
    3215: "Reproductive medicine",
    3299: "Other biomedical and clinical sciences",
    33: "Built Environment and Design",
    3301: "Architecture",
    3302: "Building",
    3303: "Design",
    3304: "Urban and regional planning",
    3399: "Other built environment and design",
    34: "Chemical Sciences",
    3401: "Analytical chemistry",
    3402: "Inorganic chemistry",
    3403: "Macromolecular and materials chemistry",
    3404: "Medicinal and biomolecular chemistry",
    3405: "Organic chemistry",
    3406: "Physical chemistry",
    3407: "Theoretical and computational chemistry",
    3499: "Other chemical sciences",
    35: "Commerce, Management, Tourism and Services",
    3501: "Accounting, auditing and accountability",
    3502: "Banking, finance and investment",
    3503: "Business systems in context",
    3504: "Commercial services",
    3505: "Human resources and industrial relations",
    3506: "Marketing",
    3507: "Strategy, management and organisational behaviour",
    3508: "Tourism",
    3509: "Transportation, logistics and supply chains",
    3599: "Other commerce, management, tourism and services",
    36: "Creative Arts and Writing",
    3601: "Art history, theory and criticism",
    3602: "Creative and professional writing",
    3603: "Music",
    3604: "Performing arts",
    3605: "Screen and digital media",
    3606: "Visual arts",
    3699: "Other creative arts and writing",
    37: "Earth Sciences",
    3701: "Atmospheric sciences",
    3702: "Climate change science",
    3703: "Geochemistry",
    3704: "Geoinformatics",
    3705: "Geology",
    3706: "Geophysics",
    3707: "Hydrology",
    3708: "Oceanography",
    3709: "Physical geography and environmental geoscience",
    3799: "Other earth sciences",
    38: "Economics",
    3801: "Applied economics",
    3802: "Econometrics",
    3803: "Economic theory",
    3899: "Other economics",
    39: "Education",
    3901: "Curriculum and pedagogy",
    3902: "Education policy, sociology and philosophy",
    3903: "Education systems",
    3904: "Specialist studies in education",
    3999: "Other education",
    40: "Engineering",
    4001: "Aerospace engineering",
    4002: "Automotive engineering",
    4003: "Biomedical engineering",
    4004: "Chemical engineering",
    4005: "Civil engineering",
    4006: "Communications engineering",
    4007: "Control engineering, mechatronics and robotics",
    4008: "Electrical engineering",
    4009: "Electronics, sensors and digital hardware",
    4010: "Engineering practice and education",
    4011: "Environmental engineering",
    4012: "Fluid mechanics and thermal engineering",
    4013: "Geomatic engineering",
    4014: "Manufacturing engineering",
    4015: "Maritime engineering",
    4016: "Materials engineering",
    4017: "Mechanical engineering",
    4018: "Nanotechnology",
    4019: "Resources engineering and extractive metallurgy",
    4099: "Other engineering",
    41: "Environmental Sciences",
    4101: "Climate change impacts and adaptation",
    4102: "Ecological applications",
    4103: "Environmental biotechnology",
    4104: "Environmental management",
    4105: "Pollution and contamination",
    4106: "Soil sciences",
    4199: "Other environmental sciences",
    42: "Health Sciences",
    4201: "Allied health and rehabilitation science",
    4202: "Epidemiology",
    4203: "Health services and systems",
    4204: "Midwifery",
    4205: "Nursing",
    4206: "Public health",
    4207: "Sports science and exercise",
    4208: "Traditional, complementary and integrative medicine",
    4299: "Other health sciences",
    43: "History, Heritage and Archaeology",
    4301: "Archaeology",
    4302: "Heritage, archive and museum studies",
    4303: "Historical studies",
    4399: "Other history, heritage and archaeology",
    44: "Human Society",
    4401: "Anthropology",
    4402: "Criminology",
    4403: "Demography",
    4404: "Development studies",
    4405: "Gender studies",
    4406: "Human geography",
    4407: "Policy and administration",
    4408: "Political science",
    4409: "Social work",
    4410: "Sociology",
    4499: "Other human society",
    45: "Indigenous Studies",
    4501: "Aboriginal and Torres Strait Islander culture, language and history",
    4502: "Aboriginal and Torres Strait Islander education",
    4503: "Aboriginal and Torres Strait Islander environmental knowledges and management",  # noqa: E501
    4504: "Aboriginal and Torres Strait Islander health and wellbeing",
    4505: "Aboriginal and Torres Strait Islander peoples, society and community",
    4506: "Aboriginal and Torres Strait Islander sciences",
    4507: "Te ahurea, reo me te hītori o te Māori (Māori culture, language and history)",  # noqa: E501
    4508: "Mātauranga Māori (Māori education)",
    4509: "Ngā mātauranga taiao o te Māori (Māori environmental knowledges)",
    4510: "Te hauora me te oranga o te Māori (Māori health and wellbeing)",
    4511: "Ngā tāngata, te porihanga me ngā hapori o te Māori (Māori peoples, society and community)",  # noqa: E501
    4512: "Ngā pūtaiao Māori (Māori sciences)",
    4513: "Pacific Peoples culture, language and history",
    4514: "Pacific Peoples education",
    4515: "Pacific Peoples environmental knowledges",
    4516: "Pacific Peoples health and wellbeing",
    4517: "Pacific Peoples sciences",
    4518: "Pacific Peoples society and community",
    4519: "Other Indigenous data, methodologies and global Indigenous studies",
    4599: "Other Indigenous studies",
    46: "Information and Computing Sciences",
    4601: "Applied computing",
    4602: "Artificial intelligence",
    4603: "Computer vision and multimedia computation",
    4604: "Cybersecurity and privacy",
    4605: "Data management and data science",
    4606: "Distributed computing and systems software",
    4607: "Graphics, augmented reality and games",
    4608: "Human-centred computing",
    4609: "Information systems",
    4610: "Library and information studies",
    4611: "Machine learning",
    4612: "Software engineering",
    4613: "Theory of computation",
    4699: "Other information and computing sciences",
    47: "Language, Communication and Culture",
    4701: "Communication and media studies",
    4702: "Cultural studies",
    4703: "Language studies",
    4704: "Linguistics",
    4705: "Literary studies",
    4799: "Other language, communication and culture",
    48: "Law and Legal Studies",
    4801: "Commercial law",
    4802: "Environmental and resources law",
    4803: "International and comparative law",
    4804: "Law in context",
    4805: "Legal systems",
    4806: "Private law and civil obligations",
    4807: "Public law",
    4899: "Other law and legal studies",
    49: "Mathematical Sciences",
    4901: "Applied mathematics",
    4902: "Mathematical physics",
    4903: "Numerical and computational mathematics",
    4904: "Pure mathematics",
    4905: "Statistics",
    4999: "Other mathematical sciences",
    50: "Philosophy and Religious Studies",
    5001: "Applied ethics",
    5002: "History and philosophy of specific fields",
    5003: "Philosophy",
    5004: "Religious studies",
    5005: "Theology",
    5099: "Other philosophy and religious studies",
    51: "Physical Sciences",
    5101: "Astronomical sciences",
    5102: "Atomic, molecular and optical physics",
    5103: "Classical physics",
    5104: "Condensed matter physics",
    5105: "Medical and biological physics",
    5106: "Nuclear and plasma physics",
    5107: "Particle and high energy physics",
    5108: "Quantum physics",
    5109: "Space sciences",
    5110: "Synchrotrons and accelerators",
    5199: "Other physical sciences",
    52: "Psychology",
    5201: "Applied and developmental psychology",
    5202: "Biological psychology",
    5203: "Clinical and health psychology",
    5204: "Cognitive and computational psychology",
    5205: "Social and personality psychology",
    5299: "Other psychology",
}
"""A list of codes to the *main* fields of researched in the ANZSRC."""


def get_name_from_code(code: int | str) -> str:
    """
    Returns:
        The full name of a Field of Research based on its code.
    """
    if isinstance(code, int) or code.isdigit():
        return ANZSRC_CLASSIFICATIONS[int(code)]

    raise ValueError(f"code is not a known classification: '{code}'")


# {{{ parse research classification


def parse_research_classification(filename: pathlib.Path) -> dict[int, str]:
    """Construct the [ANZSRC_CLASSIFICATIONS] list.

    This function takes the data from [ANZSRC_FOR_URL][] and converts it into
    a helpful dictionary.
    """
    import openpyxl

    wb = openpyxl.load_workbook(filename, read_only=True)
    if wb is None:
        raise ValueError(f"could not load workbook from file: '{filename}'")

    if "Table 2" not in wb.sheetnames:
        raise ValueError(f"expected a sheet called 'Table 2' in file: '{filename}'")
    ws = wb["Table 2"]

    # NOTE: the format of these things seems to be something like
    #
    # XX | Main category name | --
    #    | XXXX               | Subcategory name
    #
    # so we just switch on whether the first column is None or not.

    from titlecase import titlecase

    # TODO: do we want to keep the nesting in these? Not very useful for CORE
    # right now, but maybe in the future it could help..
    cls = {}
    for row in ws.iter_rows(min_row=10, max_col=3, values_only=True):
        if row[1] is None:
            break

        if row[0] is None:
            cls[int(row[1])] = row[2]
        else:
            cls[int(row[0])] = titlecase(row[1])

    return cls


# }}}
